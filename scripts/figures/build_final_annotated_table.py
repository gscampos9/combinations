"""
Standalone: reproduce the final annotated + prioritized gene-pair
supplementary workbook (comb/gene_pairs_SPARK_ys*.xlsx) directly from the
raw pipeline inputs -- RareComb outputs (results_221/) + databases/ -- so
it can be regenerated from scratch on a machine (e.g. the cluster) that
only has those two directories, without needing any of this repo's
already-computed figures/*.tsv first.

Requires databases/NDDgenes_updated.txt (the Sui_NDD686 release) to be
present in databases/ alongside the original NDDgenes.txt -- if only the
old file is there, gene1_Sui_686/gene2_Sui_686 come back all-False and
gene*_prioritized is scored on the other 3 criteria only (a warning is
printed either way).

Output: figures/Gene_pairs_SPARK_annotated.xlsx, 3 sheets (same header
names/values as gene_pairs_SPARK_ys_gsc.xlsx):
  enriched_in_pro              -- 477 case gene pairs
  enriched_in_sib              -- 1573 control gene pairs
  compare_pro-pairs_sib-pairs  -- per-gene partner-gene lists, pro vs. sib

gene1_prioritized / gene2_prioritized = NDD gene (SFARI, Fu, Wang, Zhou,
Kaplanis, Satterstrom, or Sui_686) OR DDG2P confidence definitive/strong OR
HPA brain-enriched OR pLI >= 0.9.

Usage: python build_final_annotated_table.py
"""
from __future__ import annotations

import pandas as pd

import _common as c

_PLI_CUTOFF = 0.9
_DDG2P_OK = {"definitive", "strong"}

# True stays TRUE, False is written as a blank cell instead of FALSE --
# matches the hand-built gene_pairs_SPARK_ys_gsc.xlsx's own convention.
_BOOL_COLUMNS = [
    "gene1_prioritized", "gene2_prioritized",
    "gene1_ndd_gene", "gene2_ndd_gene",
    "gene1_Sui_686", "gene2_Sui_686",
    "gene1_HPA_brain_expressed", "gene2_HPA_brain_expressed",
    "gene1_HPA_brain_enriched", "gene2_HPA_brain_enriched",
]

# Single column order applied to BOTH sheets (the two hand-edited sheets
# this replaces had drifted to slightly different orders -- standalone
# regeneration always emits this one order for both).
_PAIR_COLUMNS = [
    "Case_Samples", "Control_Samples", "combination", "gene1", "gene2", "X_chr_gene",
    "n_pro_total", "n_sib_total", "n_fam_total",
    "DNM-DNM", "DNM-maternal", "DNM-paternal",
    "maternal-paternal", "maternal-maternal", "paternal-paternal",
    "gene1_prioritized", "gene2_prioritized", "pair_STRING_ppi_score", "pair_coexpression",
    "gene1_ndd_gene", "gene1_Sui_686", "gene1_sfari_score", "gene1_DDG2P_confidence",
    "gene1_clingen_classification", "gene1_HPA_brain_expressed", "gene1_HPA_brain_enriched", "gene1_pLI",
    "gene2_ndd_gene", "gene2_Sui_686", "gene2_sfari_score", "gene2_DDG2P_confidence",
    "gene2_clingen_classification", "gene2_HPA_brain_expressed", "gene2_HPA_brain_enriched", "gene2_pLI",
]

_RENAME = {
    "n_pro": "n_pro_total", "n_sib": "n_sib_total", "n_fam": "n_fam_total",
    "inh_DD": "DNM-DNM", "inh_DM": "DNM-maternal", "inh_DP": "DNM-paternal",
    "inh_MP": "maternal-paternal", "inh_MM": "maternal-maternal", "inh_PP": "paternal-paternal",
    "ppi_score": "pair_STRING_ppi_score", "coexpressed_module": "pair_coexpression",
}
for _side in ("gene1", "gene2"):
    _RENAME[f"{_side}_g2p_dd_confidence"] = f"{_side}_DDG2P_confidence"
    _RENAME[f"{_side}_brain_expressed"] = f"{_side}_HPA_brain_expressed"
    _RENAME[f"{_side}_brain_enriched"] = f"{_side}_HPA_brain_enriched"
    _RENAME[f"{_side}_sui_hc"] = f"{_side}_Sui_686"


def _combo_key(g1, g2) -> str:
    return ",".join(sorted([g1, g2]))


def _gene_side_columns(df: pd.DataFrame, gene_col: str, db: pd.DataFrame, gene_table: pd.DataFrame, prefix: str) -> pd.DataFrame:
    genes = df[gene_col]
    db_side = db.reindex(genes).reset_index(drop=True).add_prefix(f"{prefix}_")
    gt_side = gene_table.reindex(genes).reset_index(drop=True).add_prefix(f"{prefix}_")
    return pd.concat([db_side, gt_side], axis=1)


def annotate_pairs(stats_df: pd.DataFrame, rarecomb_output_path, case_value: str,
                    ppi: dict, modules: dict) -> pd.DataFrame:
    """Only what _PAIR_COLUMNS actually needs: pair PPI/coexpression, the
    per-gene database/gnomAD columns, and inheritance-origin counts. No GO
    Jaccard/Reactome (that's build_annotations.annotate_group's job for the
    comprehensive table, not this one -- avoids needing Rgene2go_v2/Reactome
    files or the several-second GO term index build for a table that
    doesn't use them)."""
    df = stats_df.reset_index(drop=True).copy()
    df["combination"] = [_combo_key(a, b) for a, b in zip(df.gene1, df.gene2)]

    df["ppi_score"] = [c.ppi_score(g1, g2, ppi) / 1000.0 for g1, g2 in zip(df.gene1, df.gene2)]
    df["coexpressed_module"] = [c.coexpressed_pair(g1, g2, modules) for g1, g2 in zip(df.gene1, df.gene2)]

    all_genes = sorted(set(df.gene1) | set(df.gene2))
    db = c.build_database_hits(all_genes).set_index("gene")
    gene_table = c.build_gene_table(all_genes).set_index("gene")
    df = pd.concat([
        df,
        _gene_side_columns(df, "gene1", db, gene_table, "gene1"),
        _gene_side_columns(df, "gene2", db, gene_table, "gene2"),
    ], axis=1)

    print(f"  Building inheritance origin table (case_value={case_value!r}) ...")
    inh = c.build_inheritance_origin_table(rarecomb_output_path, case_value)
    inh_counts = inh.groupby("combination")[c.ORIGIN_COLS].sum().add_prefix("inh_")
    df = df.merge(inh_counts, on="combination", how="left")
    for col in c.ORIGIN_COLS:
        df[f"inh_{col}"] = df[f"inh_{col}"].fillna(0).astype(int)

    return df


def _is_prioritized(ndd, sui, ddg2p_conf, brain_enriched, pli) -> bool:
    ddg2p_hit = str(ddg2p_conf).strip().lower() in _DDG2P_OK
    pli_hit = pd.notna(pli) and float(pli) >= _PLI_CUTOFF
    return bool(ndd) or bool(sui) or ddg2p_hit or bool(brain_enriched) or pli_hit


def finalize(ann: pd.DataFrame) -> pd.DataFrame:
    df = ann.copy()

    x1 = df.get("gene1_chromosome", pd.Series("", index=df.index)).fillna("").astype(str).str.lower() == "chrx"
    x2 = df.get("gene2_chromosome", pd.Series("", index=df.index)).fillna("").astype(str).str.lower() == "chrx"
    df["X_chr_gene"] = [
        ",".join(g for g, hit in (("gene1", a), ("gene2", b)) if hit)
        for a, b in zip(x1, x2)
    ]

    for side in ("gene1", "gene2"):
        df[f"{side}_prioritized"] = [
            _is_prioritized(ndd, sui, conf, enr, pli) for ndd, sui, conf, enr, pli in
            zip(df[f"{side}_ndd_gene"], df[f"{side}_sui_hc"], df[f"{side}_g2p_dd_confidence"],
                df[f"{side}_brain_enriched"], df[f"{side}_pLI"])
        ]

    df = df.rename(columns=_RENAME)
    df = df[[col for col in _PAIR_COLUMNS if col in df.columns]]

    for col in _BOOL_COLUMNS:
        if col in df.columns:
            df[col] = df[col].map({True: True, False: None})

    return df


def compare_sheet(case_ann: pd.DataFrame, control_ann: pd.DataFrame) -> pd.DataFrame:
    """One row per gene appearing in any case or control pair: its partner
    gene(s) in pro pairs and in sib pairs (comma-joined, alphabetical)."""
    def partners(df):
        out = {}
        for g1, g2 in zip(df["gene1"], df["gene2"]):
            out.setdefault(g1, set()).add(g2)
            out.setdefault(g2, set()).add(g1)
        return out

    pro_partners = partners(case_ann)
    sib_partners = partners(control_ann)
    all_genes = sorted(set(pro_partners) | set(sib_partners))

    return pd.DataFrame({
        "gene": all_genes,
        "paired gene in pro (one)": [
            ",".join(sorted(pro_partners.get(g, ()))) or None for g in all_genes],
        "paired gene in sib (one)": [
            ",".join(sorted(sib_partners.get(g, ()))) or None for g in all_genes],
    })


def _apply_number_formats(path):
    """Uniform display format for every pLI / PPI-score column, so the
    workbook never drifts into the mixed Scientific/General formatting a
    copy-paste build-up produces. Plain decimal, not scientific -- Excel's
    own "General" format silently switches to scientific notation for very
    small numbers (pLI can be as small as 1e-226), so an explicit fixed
    format is required; values that small still display as a string of
    zeros (double precision only carries ~15-17 significant digits, and
    Excel's own display caps out around there too), but the exact stored
    number is unaffected -- only how it's shown."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        header = {cell.value: cell.column for cell in ws[1]}
        for col_name, fmt in (
            ("gene1_pLI", "0." + "0" * 20), ("gene2_pLI", "0." + "0" * 20),
            ("pair_STRING_ppi_score", "0.000"),
        ):
            col = header.get(col_name)
            if col is None:
                continue
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = fmt
    wb.save(path)


def main():
    case_df = c.load_stats(str(c.CASE_OUTPUT))
    control_df = c.load_stats(str(c.CONTROL_OUTPUT))

    all_genes = sorted(set(case_df.gene1) | set(case_df.gene2) | set(control_df.gene1) | set(control_df.gene2))
    ppi = c.load_ppi_for_genes(all_genes)
    modules = c.load_coexpression_modules()
    # No GO/Reactome, no permutation results: not in this table's column set
    # (_PAIR_COLUMNS) so they're never loaded at all, not even to discard
    # afterward -- keeps this to exactly the headers that already existed.

    c._load_database_sets()  # noqa: SLF001 -- runs first so the "NDD evidence: ... from <file>" line prints up front

    print("Annotating case gene pairs (probands) ...")
    case_ann = finalize(annotate_pairs(case_df, c.CASE_OUTPUT, "asd", ppi, modules))
    print(f"  {len(case_ann)} pairs, "
          f"{int((case_ann.gene1_prioritized.notna() | case_ann.gene2_prioritized.notna()).sum())} "
          f"with >=1 prioritized gene")

    print("Annotating control gene pairs (siblings) ...")
    control_ann = finalize(annotate_pairs(control_df, c.CONTROL_OUTPUT, "sib", ppi, modules))
    print(f"  {len(control_ann)} pairs, "
          f"{int((control_ann.gene1_prioritized.notna() | control_ann.gene2_prioritized.notna()).sum())} "
          f"with >=1 prioritized gene")

    print("Building compare_pro-pairs_sib-pairs sheet ...")
    compare_df = compare_sheet(case_ann, control_ann)
    print(f"  {len(compare_df)} genes")

    out_path = c.FIGURES_DIR / "Gene_pairs_SPARK_annotated.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        case_ann.to_excel(xl, sheet_name="enriched_in_pro", index=False)
        control_ann.to_excel(xl, sheet_name="enriched_in_sib", index=False)
        compare_df.to_excel(xl, sheet_name="compare_pro-pairs_sib-pairs", index=False)
    _apply_number_formats(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
